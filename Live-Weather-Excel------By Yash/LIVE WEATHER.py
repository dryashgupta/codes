import requests
import openpyxl
import time

wb = openpyxl.load_workbook('test1.xlsx')
sheet_obj = wb.active

api_key = '4e41613b287a52b6be9b1176c338eab9'          #Enter your api key if you have here

count = 0
while True:
    rows = sheet_obj.max_row

    for i in range(2, rows+1):
        city = sheet_obj.cell(row = i, column = 1).value
        temp = sheet_obj.cell(row = i, column = 2).value
        humidity = sheet_obj.cell(row = i , column = 3).value
        unit = input('Enter the Temperature Unit C/F ')
        update = sheet_obj.cell(row = i, column = 5).value
        #print(city,temp,humidity,unit,update)
        if count == 0:
            if unit.upper() == 'F':
                unit = 'imperial'
            else:
                unit = 'metric'
            data = requests.get('http://api.openweathermap.org/data/2.5/weather?apikey={}&mode=json&units={}&q={}'.format(api_key,unit,city)).json()
            #print(data)
            temp = data['main']['temp']
            humidity = data['main']['humidity']
            sheet_obj.cell(column=2 , row=i, value=temp)
            sheet_obj.cell(column=3 , row=i, value=humidity)
        elif count!=0 and update == 1:
            if unit == 'F':
                unit = 'imperial'
            else:
                unit = 'metric'
            data = requests.get('http://api.openweathermap.org/data/2.5/weather?apikey={}&mode=json&units={}&q={}'.format(api_key,unit,city)).json()
            #print(data)
            temp = data['main']['temp']
            humidity = data['main']['humidity']
            sheet_obj.cell(column=2 , row=i, value=temp)
            sheet_obj.cell(column=3 , row=i, value=humidity)
            sheet_obj.cell(column=4 , row=i, value=unit)
        else:
            continue
    wb.save('test1.xlsx')
    print("Info Updated...")
    count+=1
    time.sleep(5)
    x = input('Do u want to continue y/n ')
    if x.lower() == 'y':
        continue
    else:
        break
#File gets updated every 5 second.
